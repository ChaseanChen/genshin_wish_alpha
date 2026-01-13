from openpyxl import load_workbook

def get_data_from_excel(file_path, sheet_name, target_value_col3):
    """
    在Excel的第三列查找指定值，并返回对应的第一列和第二列数据。

    Args:
        file_path (str): Excel文件的路径。
        sheet_name (str): 要处理的工作表名称。
        target_value_col3: 要在第三列查找的固定数值。

    Returns:
        list: 包含字典的列表，每个字典包含匹配行的第一列和第二列数据。
              例如：[{'第一列': 'Apple', '第二列': 'Red'}, {'第一列': 'Date', '第二列': 'Brown'}]
    """
    results = []
    try:
        # 加载工作簿
        workbook = load_workbook(filename=file_path)
        # 选择工作表
        sheet = workbook[sheet_name]

        # 遍历工作表中的每一行
        # sheet.iter_rows() 返回的是单元格对象，我们可以通过 .value 获取其值
        # min_row=2 表示从第二行开始读取，跳过标题行。如果你的数据没有标题行，请设置为 1。
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2): # start=2 for actual row number in excel
            # 检查行是否有足够的列（至少3列）
            if len(row) < 3:
                continue

            # 获取第三列的单元格值 (Python的列表索引是0开始的，所以第三列是索引2)
            col3_value = row[2].value

            # 将目标值和单元格值都转换为字符串进行比较，增加鲁棒性
            # 或者确保它们的数据类型一致，特别是当Excel单元格格式不确定时
            if str(col3_value) == str(target_value_col3):
                # 获取第一列和第二列的单元格值 (索引0和1)
                col1_value = row[0].value
                col2_value = row[1].value

                results.append({
                    "行号": row_index, # 可以选择包含行号
                    "第一列": col1_value,
                    "第二列": col2_value
                })

    except FileNotFoundError:
        print(f"错误：文件 '{file_path}' 未找到。")
    except KeyError:
        print(f"错误：工作表 '{sheet_name}' 未找到。")
    except Exception as e:
        print(f"发生未知错误: {e}")

    return results

# --- 使用示例 ---
excel_file = 'your_excel_file.xlsx'  # 替换为你的Excel文件路径
sheet_name = 'Sheet1'                 # 替换为你的工作表名称
target_number = 100                   # 替换为你要查找的固定数值

found_data = get_data_from_excel(excel_file, sheet_name, target_number)

if found_data:
    print(f"在第三列找到 '{target_number}' 的数据：")
    for item in found_data:
        print(f"  行号: {item['行号']}, 第一列: {item['第一列']}, 第二列: {item['第二列']}")
else:
    print(f"在第三列中未找到 '{target_number}'。")